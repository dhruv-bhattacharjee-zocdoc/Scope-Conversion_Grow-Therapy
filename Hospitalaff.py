import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

# Path to the output Excel file
output_file = 'Excel Files/Output.xlsx'
provider_sheet_name = 'Provider'
validation_sheet_name = 'ValidationAndReference'
validation_range = 'ValidationAndReference!$T$2:$T$7258'

# Columns to apply validation to
affiliation_columns = [
    'Hospital Affiliation 1',
    'Hospital Affiliation 2',
    'Hospital Affiliation 3',
    'Hospital Affiliation 4',
    'Hospital Affiliation 5',
]

def main():
    wb = openpyxl.load_workbook(output_file)
    ws = wb[provider_sheet_name]

    # Find the column indices for the target columns
    header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    col_indices = []
    for col_name in affiliation_columns:
        try:
            col_idx = header.index(col_name) + 1  # openpyxl is 1-indexed
            col_indices.append(col_idx)
        except ValueError:
            print(f"Column '{col_name}' not found in header.")
            continue

    # Determine the last row with data
    max_row = ws.max_row

    # Create the data validation object
    dv = DataValidation(type="list", formula1=f'={validation_range}', allow_blank=True)
    dv.error = 'Select a value from the dropdown.'
    dv.errorTitle = 'Invalid Entry'
    dv.prompt = 'Please select a hospital affiliation from the list.'
    dv.promptTitle = 'Hospital Affiliation'

    # Add the data validation to the worksheet
    ws.add_data_validation(dv)

    # Apply the validation to each relevant column (excluding header row)
    for col_idx in col_indices:
        col_letter = get_column_letter(col_idx)
        dv.add(f'{col_letter}2:{col_letter}{max_row}')

    # Save the workbook
    wb.save(output_file)
    print('Data validation added successfully.')

if __name__ == '__main__':
    main()
