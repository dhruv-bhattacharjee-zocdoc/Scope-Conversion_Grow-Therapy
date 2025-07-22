import openpyxl
import os
from openpyxl.styles import PatternFill

def copy_headshot_column(input_file, output_file, input_sheet_name=None, output_sheet_name='Provider'):
    """
    Copy values from the 'Latest Headshot' column in the input Excel file to the 'Headshot Link' column in the 'Provider' sheet of the output Excel file.
    Fill empty cells in the 'Headshot Link' column with the color FFFFC7CE.
    """
    wb_in = openpyxl.load_workbook(input_file)
    wb_out = openpyxl.load_workbook(output_file)

    ws_in = wb_in[input_sheet_name] if input_sheet_name else wb_in.worksheets[0]
    ws_out = wb_out[output_sheet_name]

    # Find the column index for 'Latest Headshot' in input and 'Headshot Link' in output
    in_header = [cell.value for cell in next(ws_in.iter_rows(min_row=1, max_row=1))]
    out_header = [cell.value for cell in next(ws_out.iter_rows(min_row=1, max_row=1))]

    try:
        headshot_col_in = in_header.index('Latest Headshot') + 1  # 1-based index
        headshot_col_out = out_header.index('Headshot Link') + 1
    except ValueError as e:
        raise Exception(f"Column not found: {e}")

    # Copy values from input to output (starting from row 2)
    input_data_rows = 0
    for row_idx, row in enumerate(ws_in.iter_rows(min_row=2), start=2):
        headshot_value = row[headshot_col_in - 1].value
        ws_out.cell(row=row_idx, column=headshot_col_out, value=headshot_value)
        input_data_rows += 1

    # Fill empty cells in 'Headshot Link' column with FFFFC7CE for all rows in the raw table
    fill_color = PatternFill(start_color='FFFFC7CE', end_color='FFFFC7CE', fill_type='solid')
    for row_idx in range(2, 2 + input_data_rows):
        cell = ws_out.cell(row=row_idx, column=headshot_col_out)
        if cell.value is None or cell.value == "":
            cell.fill = fill_color

    wb_out.save(output_file)

if __name__ == '__main__':
    input_sample_path = os.path.join('Excel Files', 'Grow Therapy - Sample Data.xlsx')
    output_path = os.path.join('Excel Files', 'Output.xlsx')
    copy_headshot_column(input_sample_path, output_path, output_sheet_name='Provider')
