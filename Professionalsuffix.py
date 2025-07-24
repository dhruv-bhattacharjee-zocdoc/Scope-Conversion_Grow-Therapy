import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
import os
from openpyxl.styles import PatternFill
from difflib import get_close_matches

def copy_professional_suffix(input_file, output_file, input_sheet_name=None, output_sheet_name='Provider'):
    """
    Copy values from the 'Board Certification' column in the input Excel file to the 'Professional Suffix 1' column in the 'Provider' sheet of the output Excel file.
    Add a dropdown to 'Professional Suffix 1' using '=ValidationAndReference!$G$2:$G$511'.
    """
    wb_in = openpyxl.load_workbook(input_file)
    wb_out = openpyxl.load_workbook(output_file)

    ws_in = wb_in[input_sheet_name] if input_sheet_name else wb_in.worksheets[0]
    ws_out = wb_out[output_sheet_name]

    # Find the column indexes
    in_header = [cell.value for cell in next(ws_in.iter_rows(min_row=1, max_row=1))]
    out_header = [cell.value for cell in next(ws_out.iter_rows(min_row=1, max_row=1))]

    try:
        board_cert_col_in = in_header.index('Board Certification') + 1  # 1-based index
        prof_suffix_col_out = out_header.index('Professional Suffix 1') + 1
        # Add for Suffix 2 and 3
        prof_suffix2_col_out = out_header.index('Professional Suffix 2') + 1
        prof_suffix3_col_out = out_header.index('Professional Suffix 3') + 1
    except ValueError as e:
        raise Exception(f"Column not found: {e}")

    # Load dropdown options from ValidationAndReference!$G$2:$G$511
    if 'ValidationAndReference' not in wb_out.sheetnames:
        raise Exception("'ValidationAndReference' sheet not found in output file.")
    ws_validation = wb_out['ValidationAndReference']
    dropdown_options = [ws_validation[f'G{row}'].value for row in range(2, 512)]
    dropdown_options = [opt for opt in dropdown_options if opt is not None and str(opt).strip() != '']

    yellow_fill = PatternFill(start_color='FFFFFF00', end_color='FFFFFF00', fill_type='solid')

    max_row = 1
    for row_idx, row in enumerate(ws_in.iter_rows(min_row=2), start=2):
        board_cert_value = row[board_cert_col_in - 1].value
        if board_cert_value is not None and str(board_cert_value).strip() != "":
            value = str(board_cert_value).strip()
            # Try exact match first
            if value in dropdown_options:
                ws_out.cell(row=row_idx, column=prof_suffix_col_out, value=value)
            else:
                # Fuzzy match to nearest option
                matches = get_close_matches(value, dropdown_options, n=1, cutoff=0.6)
                if matches:
                    ws_out.cell(row=row_idx, column=prof_suffix_col_out, value=matches[0])
                    ws_out.cell(row=row_idx, column=prof_suffix_col_out).fill = yellow_fill
                else:
                    ws_out.cell(row=row_idx, column=prof_suffix_col_out, value=value)
                    ws_out.cell(row=row_idx, column=prof_suffix_col_out).fill = yellow_fill
        max_row = row_idx
        # If the cell is empty, leave as is to preserve dropdown/data validation

    # Add dropdown to 'Professional Suffix 1', 'Professional Suffix 2', and 'Professional Suffix 3' for all relevant rows
    dv1 = DataValidation(type="list", formula1="=ValidationAndReference!$G$2:$G$511", allow_blank=True)
    dv2 = DataValidation(type="list", formula1="=ValidationAndReference!$G$2:$G$511", allow_blank=True)
    dv3 = DataValidation(type="list", formula1="=ValidationAndReference!$G$2:$G$511", allow_blank=True)
    col_letter1 = get_column_letter(prof_suffix_col_out)
    col_letter2 = get_column_letter(prof_suffix2_col_out)
    col_letter3 = get_column_letter(prof_suffix3_col_out)
    dv_range1 = f"{col_letter1}2:{col_letter1}{max_row}"
    dv_range2 = f"{col_letter2}2:{col_letter2}{max_row}"
    dv_range3 = f"{col_letter3}2:{col_letter3}{max_row}"
    dv1.add(dv_range1)
    dv2.add(dv_range2)
    dv3.add(dv_range3)
    ws_out.add_data_validation(dv1)
    ws_out.add_data_validation(dv2)
    ws_out.add_data_validation(dv3)

    # Find the column index for 'Professional Suffix ID 1'
    try:
        prof_suffix_id_col_out = out_header.index('Professional Suffix ID 1') + 1
    except ValueError as e:
        raise Exception(f"Column not found: {e}")
    prof_suffix_col_letter = get_column_letter(prof_suffix_col_out)
    prof_suffix_id_col_letter = get_column_letter(prof_suffix_id_col_out)

    for row_idx in range(2, max_row + 1):
        formula = f'=IFERROR(XLOOKUP({prof_suffix_col_letter}{row_idx}, ValidationAndReference!G:G, ValidationAndReference!F:F, ""), "")'
        ws_out.cell(row=row_idx, column=prof_suffix_id_col_out, value=formula)

    wb_out.save(output_file)

if __name__ == '__main__':
    input_sample_path = os.path.join('Excel Files', 'Grow Therapy - Sample Data.xlsx')
    output_path = os.path.join('Excel Files', 'Output.xlsx')
    copy_professional_suffix(input_sample_path, output_path, output_sheet_name='Provider')
