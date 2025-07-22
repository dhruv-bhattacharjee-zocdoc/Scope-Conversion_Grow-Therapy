import openpyxl
import os
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

def copy_languages(input_file, output_file, input_sheet_name=None, output_sheet_name='Provider'):
    """
    Copy and split 'Latest Languages' from the input file into 'Additional Languages Spoken 1/2/3' in the Provider sheet of the output file.
    Preserve dropdowns/data validation in the output columns.
    Add a dropdown to 'Additional Languages Spoken 1/2/3' using '=ValidationAndReference!$W$2:$W$144'.
    """
    wb_in = openpyxl.load_workbook(input_file)
    wb_out = openpyxl.load_workbook(output_file)

    ws_in = wb_in[input_sheet_name] if input_sheet_name else wb_in.worksheets[0]
    ws_out = wb_out[output_sheet_name]

    # Find the column indexes
    in_header = [cell.value for cell in next(ws_in.iter_rows(min_row=1, max_row=1))]
    out_header = [cell.value for cell in next(ws_out.iter_rows(min_row=1, max_row=1))]

    try:
        lang_col_in = in_header.index('Latest Languages') + 1  # 1-based index
        lang1_col_out = out_header.index('Additional Languages Spoken 1') + 1
        lang2_col_out = out_header.index('Additional Languages Spoken 2') + 1
        lang3_col_out = out_header.index('Additional Languages Spoken 3') + 1
    except ValueError as e:
        raise Exception(f"Column not found: {e}")

    # Copy and split values from input to output (starting from row 2)
    max_row = 1
    for row_idx, row in enumerate(ws_in.iter_rows(min_row=2), start=2):
        lang_value = row[lang_col_in - 1].value
        if lang_value is not None and str(lang_value).strip() != "":
            lang_str = str(lang_value)
            langs = [l.strip() for l in lang_str.replace('+', ' ').split() if l.strip()]
            if len(langs) > 0:
                ws_out.cell(row=row_idx, column=lang1_col_out, value=langs[0])
            if len(langs) > 1:
                ws_out.cell(row=row_idx, column=lang2_col_out, value=langs[1])
            if len(langs) > 2:
                ws_out.cell(row=row_idx, column=lang3_col_out, value=langs[2])
        max_row = row_idx
        # If the cell is empty, leave as is to preserve dropdown/data validation

    # Add dropdown to 'Additional Languages Spoken 1/2/3' for all relevant rows
    for col_out in [lang1_col_out, lang2_col_out, lang3_col_out]:
        dv = DataValidation(type="list", formula1="=ValidationAndReference!$W$2:$W$144", allow_blank=True)
        col_letter = get_column_letter(col_out)
        dv_range = f"{col_letter}2:{col_letter}{max_row}"
        dv.add(dv_range)
        ws_out.add_data_validation(dv)

    wb_out.save(output_file) 