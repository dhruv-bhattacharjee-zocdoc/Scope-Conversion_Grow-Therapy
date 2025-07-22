import openpyxl
import os

def copy_patients_accepted(input_file, output_file, input_sheet_name=None, output_sheet_name='Provider'):
    """
    Copy and process the 'Latest age_focus' column from the input Excel file to the 'Patients Accepted' column in the 'Provider' sheet of the output Excel file.
    Map the value to 'Adult', 'Pediatric', or 'Both' based on the presence of age groups.
    """
    wb_in = openpyxl.load_workbook(input_file)
    wb_out = openpyxl.load_workbook(output_file)

    ws_in = wb_in[input_sheet_name] if input_sheet_name else wb_in.worksheets[0]
    ws_out = wb_out[output_sheet_name]

    # Find the column indexes
    in_header = [cell.value for cell in next(ws_in.iter_rows(min_row=1, max_row=1))]
    out_header = [cell.value for cell in next(ws_out.iter_rows(min_row=1, max_row=1))]

    try:
        age_focus_col_in = in_header.index('Latest age_focus') + 1  # 1-based index
        patients_accepted_col_out = out_header.index('Patients Accepted') + 1
    except ValueError as e:
        raise Exception(f"Column not found: {e}")

    # Define age group mappings
    adult_terms = {'Adults (18 to 64)', 'Elders (65 and above)'}
    pediatric_terms = {'Children (6 to 12)', 'Teenagers (13 to 17)'}

    for row_idx, row in enumerate(ws_in.iter_rows(min_row=2), start=2):
        age_focus_value = row[age_focus_col_in - 1].value
        result = ''
        if age_focus_value is not None and str(age_focus_value).strip() != '':
            parts = [p.strip() for p in str(age_focus_value).split('+')]
            has_adult = any(part in adult_terms for part in parts)
            has_pediatric = any(part in pediatric_terms for part in parts)
            if has_adult and has_pediatric:
                result = 'Both'
            elif has_adult:
                result = 'Adult'
            elif has_pediatric:
                result = 'Pediatric'
            else:
                result = ''  # Or 'Unknown' if you want to flag it
        ws_out.cell(row=row_idx, column=patients_accepted_col_out, value=result)

    wb_out.save(output_file)
