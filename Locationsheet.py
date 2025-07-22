import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
import re

def process_location_sheet(input_path, output_path):
    # Load workbooks
    input_wb = openpyxl.load_workbook(input_path)
    output_wb = openpyxl.load_workbook(output_path)

    # Assume only one sheet in input
    input_ws = input_wb.active
    location_ws = output_wb['Location']
    validation_ws = output_wb['ValidationAndReference']

    # Get headers from input
    input_headers = {cell.value: idx for idx, cell in enumerate(next(input_ws.iter_rows(min_row=1, max_row=1)), 1)}

    # Output headers (assume first row)
    output_headers = {cell.value: idx for idx, cell in enumerate(next(location_ws.iter_rows(min_row=1, max_row=1)), 1)}

    # Helper to get column letter by header name
    def col_letter(ws, header):
        idx = output_headers[header]
        return openpyxl.utils.get_column_letter(idx)

    # Data validation setup
    def add_dropdown(ws, col, options, first_row, last_row):
        dv = DataValidation(type="list", formula1=options, allow_blank=True)
        ws.add_data_validation(dv)
        dv.add(f"{col}{first_row}:{col}{last_row}")

    def add_dropdown_values(ws, col, values, first_row, last_row):
        options = '"' + ','.join(values) + '"'
        add_dropdown(ws, col, options, first_row, last_row)

    # Find data range
    first_data_row = 2
    last_data_row = input_ws.max_row

    # Map and copy data
    for row in range(first_data_row, last_data_row + 1):
        out_row = row
        # Phone Number -> Phone
        phone = input_ws.cell(row=row, column=input_headers['Phone Number']).value
        location_ws.cell(row=out_row, column=output_headers['Phone']).value = phone
        # Email addresses to send notifcations to (could be same for all) -> Email for appointment notifications 1
        email = input_ws.cell(row=row, column=input_headers['Email addresses to send notifcations to (could be same for all)']).value
        location_ws.cell(row=out_row, column=output_headers['Email for appointment notifications 1']).value = email
        # Extract ZIP Code from address
        address = input_ws.cell(row=row, column=input_headers['Primary Treating Clinic Location 1 (Clinic name and address)']).value
        zip_code = None
        if address:
            match = re.search(r'(\d{5})', address)
            if match:
                zip_code = match.group(1)
        location_ws.cell(row=out_row, column=output_headers['ZIP Code']).value = zip_code
        # Extract State from address (robust: before ZIP or after last comma)
        state = None
        if address:
            # Find all two-letter uppercase abbreviations after a comma
            matches = re.findall(r',\s*([A-Z]{2})\b', address)
            if matches:
                state = matches[-1]  # Use the last match
        location_ws.cell(row=out_row, column=output_headers['State']).value = state
        # Improved City extraction using regex
        city = None
        if address:
            # Look for ', City, STATE' or ', City, STATE ZIP'
            match = re.search(r',\s*([^,]+),\s*[A-Z]{2}\b', address)
            if match:
                city = match.group(1).strip()
        location_ws.cell(row=out_row, column=output_headers['City']).value = city
        # Name -> Name of Contact Person
        contact_name = input_ws.cell(row=row, column=input_headers['Name']).value
        location_ws.cell(row=out_row, column=output_headers['Name of Contact Person']).value = contact_name
        # Scheduling Software ID formula
        sched_col = output_headers['Scheduling Software']
        sched_id_col = output_headers['Scheduling Software ID']
        sched_col_letter = openpyxl.utils.get_column_letter(sched_col)
        sched_id_col_letter = openpyxl.utils.get_column_letter(sched_id_col)
        formula = f'=IF(ISBLANK({sched_col_letter}{out_row}),"",INDEX(ValidationAndReference!C:C,MATCH({sched_col_letter}{out_row},ValidationAndReference!D:D,0)))'
        location_ws.cell(row=out_row, column=sched_id_col).value = formula

    # Add dropdowns
    # Location Type: In Person, Virtual
    drop_col = col_letter(location_ws, 'Location Type')
    add_dropdown_values(location_ws, drop_col, ['In Person', 'Virtual'], first_data_row, last_data_row)

    # Virtual Visit Type: ValidationAndReference!$Y$2:$Y$3
    drop_col = col_letter(location_ws, 'Virtual Visit Type')
    add_dropdown(location_ws, drop_col, '=ValidationAndReference!$Y$2:$Y$3', first_data_row, last_data_row)

    # Show name in search?: Yes, No
    drop_col = col_letter(location_ws, 'Show name in search?')
    add_dropdown_values(location_ws, drop_col, ['Yes', 'No'], first_data_row, last_data_row)

    # Scheduling Software: ValidationAndReference!$D$2:$D$750
    drop_col = col_letter(location_ws, 'Scheduling Software')
    add_dropdown(location_ws, drop_col, '=ValidationAndReference!$D$2:$D$750', first_data_row, last_data_row)

    # State: ValidationAndReference!$A$2:$A$55
    drop_col = col_letter(location_ws, 'State')
    add_dropdown(location_ws, drop_col, '=ValidationAndReference!$A$2:$A$55', first_data_row, last_data_row)

    # Save output
    output_wb.save(output_path)
