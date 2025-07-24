import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
import os
from API import run_api

# Paths
input_path = os.path.join('Excel Files', 'New Business Scope Sheet - Practice Locations and Providers.xlsx')
output_path = os.path.join('Excel Files', 'Output.xlsx')

def copy_worksheet_values_and_validations(src_ws, tgt_ws):
    # Copy values only (no formatting)
    for row in src_ws.iter_rows():
        for cell in row:
            tgt_ws[cell.coordinate].value = cell.value
    # Copy data validations
    if src_ws.data_validations:
        for dv in src_ws.data_validations.dataValidation:
            tgt_ws.add_data_validation(dv)
    # Copy merged cells
    for merged_range in src_ws.merged_cells.ranges:
        tgt_ws.merge_cells(str(merged_range))

def main():
    # Load the source workbook
    wb_src = openpyxl.load_workbook(input_path, data_only=False)
    # Create a new workbook for output
    wb_tgt = openpyxl.Workbook()
    # Remove the default sheet if it exists
    if wb_tgt.active and wb_tgt.active.title in wb_tgt.sheetnames:
        wb_tgt.remove(wb_tgt[wb_tgt.active.title])
    # Copy each sheet
    for sheet_name in wb_src.sheetnames:
        src_ws = wb_src[sheet_name]
        tgt_ws = wb_tgt.create_sheet(title=sheet_name)
        copy_worksheet_values_and_validations(src_ws, tgt_ws)
    # Copy named ranges
    for named_range in wb_src.defined_names.values():
        wb_tgt.defined_names.append(named_range)
    # Save the output workbook
    wb_tgt.save(output_path)
    print(f'Copied to {output_path} (values, data validations, named ranges, no formatting)')
    # os.startfile(output_path)  # Moved to end of script

def write_provider_type_substatus_id_formula(output_path, sheet_name='Provider'):
    import openpyxl
    wb = openpyxl.load_workbook(output_path)
    ws = wb[sheet_name]
    max_row = ws.max_row
    col_idx = 61  # 1-based index for 'Provider Type (Substatus) ID'
    assoc_col_idx = 57  # 1-based index for 'Provider Type (Substatus)' (column BE)
    for row in range(2, max_row + 1):
        assoc_value = ws.cell(row=row, column=assoc_col_idx).value
        if assoc_value is not None and str(assoc_value).strip() != '':
            formula = f'=IFERROR(XLOOKUP(BE{row}, ValidationAndReference!Q:Q, ValidationAndReference!P:P, ""), "")'
            ws.cell(row=row, column=col_idx).value = formula
        else:
            ws.cell(row=row, column=col_idx).value = ''
    wb.save(output_path)

def add_board_certification_dropdowns(output_path, sheet_name='Provider'):
    from openpyxl.utils import get_column_letter
    wb = openpyxl.load_workbook(output_path)
    ws = wb[sheet_name]
    header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    board_cert_cols = []
    for i in range(1, 6):
        col_name = f'Board Certification {i}'
        try:
            col_idx = header.index(col_name) + 1  # openpyxl is 1-indexed
            board_cert_cols.append(col_idx)
        except ValueError:
            print(f"Column '{col_name}' not found in header.")
            continue
    max_row = ws.max_row
    from openpyxl.worksheet.datavalidation import DataValidation
    dv = DataValidation(type="list", formula1="=ValidationAndReference!$N$2:$N$299", allow_blank=True)
    dv.error = 'Select a value from the dropdown.'
    dv.errorTitle = 'Invalid Entry'
    dv.prompt = 'Please select a board certification from the list.'
    dv.promptTitle = 'Board Certification'
    ws.add_data_validation(dv)
    for col_idx in board_cert_cols:
        col_letter = get_column_letter(col_idx)
        dv.add(f'{col_letter}2:{col_letter}{max_row}')
    wb.save(output_path)

def add_sub_board_certification_dropdowns(output_path, sheet_name='Provider'):
    from openpyxl.utils import get_column_letter
    wb = openpyxl.load_workbook(output_path)
    ws = wb[sheet_name]
    header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    sub_board_cert_cols = []
    for i in range(1, 6):
        col_name = f'Sub Board Certification {i}'
        try:
            col_idx = header.index(col_name) + 1  # openpyxl is 1-indexed
            sub_board_cert_cols.append(col_idx)
        except ValueError:
            print(f"Column '{col_name}' not found in header.")
            continue
    max_row = ws.max_row
    from openpyxl.worksheet.datavalidation import DataValidation
    dv = DataValidation(type="list", formula1="=ValidationAndReference!$N$2:$N$294", allow_blank=True)
    dv.error = 'Select a value from the dropdown.'
    dv.errorTitle = 'Invalid Entry'
    dv.prompt = 'Please select a sub board certification from the list.'
    dv.promptTitle = 'Sub Board Certification'
    ws.add_data_validation(dv)
    for col_idx in sub_board_cert_cols:
        col_letter = get_column_letter(col_idx)
        dv.add(f'{col_letter}2:{col_letter}{max_row}')
    wb.save(output_path)

def write_board_cert_id_1_formula(output_path, sheet_name='Provider'):
    import openpyxl
    wb = openpyxl.load_workbook(output_path)
    ws = wb[sheet_name]
    header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    try:
        board_cert_id_col = header.index('Board Cert ID 1') + 1  # 1-based index
        board_cert_col = header.index('Board Certification 1') + 1  # 1-based index (AA)
    except ValueError as e:
        print(f"Column not found: {e}")
        return
    max_row = ws.max_row
    from openpyxl.utils import get_column_letter
    board_cert_col_letter = get_column_letter(board_cert_col)
    for row in range(2, max_row + 1):
        formula = f'=IF(ISBLANK({board_cert_col_letter}{row}),"",INDEX(ValidationAndReference!$AA:$AA,MATCH({board_cert_col_letter}{row},ValidationAndReference!$AB:$AB,0)))'
        ws.cell(row=row, column=board_cert_id_col).value = formula
    wb.save(output_path)

def write_professional_suffix_id_1_formula(output_path, sheet_name='Provider'):
    import openpyxl
    wb = openpyxl.load_workbook(output_path)
    ws = wb[sheet_name]
    header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    try:
        suffix_id_col = header.index('Professional Suffix ID 1') + 1  # 1-based index
        d_col_letter = 'D'  # Column D is the lookup value
    except ValueError as e:
        print(f"Column not found: {e}")
        return
    max_row = ws.max_row
    for row in range(2, max_row + 1):
        formula = f'=IFERROR(@XLOOKUP({d_col_letter}{row}, ValidationAndReference!G:G, ValidationAndReference!F:F, ""), "")'
        ws.cell(row=row, column=suffix_id_col).value = formula
    wb.save(output_path)

def copy_location_ids_to_provider(input_file, output_file, input_sheet_name=None, output_sheet_name='Provider'):
    """
    Copy 'CORRECT Location' to 'Location ID 1', and 'Location ID 2-5' to 'Location ID 2-5' from input to output Provider sheet.
    If any column or cell is missing, leave the output cell blank.
    Highlight columns 'Location ID 1' through 'Location ID 5' for any row where 'Location ID 5' has a value.
    Only keep as many rows in the output as there are data rows in the input. Print the number of providers to the terminal.
    """
    import openpyxl
    from openpyxl.styles import PatternFill
    wb_in = openpyxl.load_workbook(input_file)
    wb_out = openpyxl.load_workbook(output_file)

    ws_in = wb_in[input_sheet_name] if input_sheet_name else wb_in.worksheets[0]
    ws_out = wb_out[output_sheet_name]

    # Define mapping: input column -> output column
    col_map = {
        'CORRECT Location': 'Location ID 1',
        'Location ID 2': 'Location ID 2',
        'Location ID 3': 'Location ID 3',
        'Location ID 4': 'Location ID 4',
        'Location ID 5': 'Location ID 5',
    }

    # Get headers
    in_header = [cell.value for cell in next(ws_in.iter_rows(min_row=1, max_row=1))]
    out_header = [cell.value for cell in next(ws_out.iter_rows(min_row=1, max_row=1))]

    # Get column indices (1-based)
    in_col_indices = {k: (in_header.index(k) + 1 if k in in_header else None) for k in col_map.keys()}
    out_col_indices = {v: (out_header.index(v) + 1 if v in out_header else None) for v in col_map.values()}

    # Prepare highlight fill (yellow)
    highlight_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

    # Determine number of data rows in input (excluding header)
    input_data_rows = ws_in.max_row - 1
    print(f"Number of providers in the batch: {input_data_rows}")

    # Copy values row by row (starting from row 2)
    for row_idx in range(2, ws_in.max_row + 1):
        # Track if Location ID 5 has a value for this row
        location_id_5_value = None
        for in_col, out_col in col_map.items():
            in_idx = in_col_indices[in_col]
            out_idx = out_col_indices[out_col]
            value = None
            if in_idx is not None and out_idx is not None:
                value = ws_in.cell(row=row_idx, column=in_idx).value
                ws_out.cell(row=row_idx, column=out_idx, value=value if value is not None else None)
            elif out_idx is not None:
                ws_out.cell(row=row_idx, column=out_idx, value=None)
            if out_col == 'Location ID 5':
                location_id_5_value = value
        # If Location ID 5 has a value, highlight Location ID 1-5 for this row
        if location_id_5_value not in [None, '']:
            for out_col in ['Location ID 1', 'Location ID 2', 'Location ID 3', 'Location ID 4', 'Location ID 5']:
                out_idx = out_col_indices[out_col]
                if out_idx is not None:
                    ws_out.cell(row=row_idx, column=out_idx).fill = highlight_fill

    # Remove extra rows from output Provider sheet
    output_data_rows = ws_out.max_row - 1
    if output_data_rows > input_data_rows:
        ws_out.delete_rows(input_data_rows + 2, output_data_rows - input_data_rows)

    wb_out.save(output_file)

if __name__ == '__main__':
    run_api()
    main()
    # Copy location IDs to Provider sheet
    input_sample_path = os.path.join('Excel Files', 'Grow Therapy - Sample Data.xlsx')
    output_path = os.path.join('Excel Files', 'Output.xlsx')
    copy_location_ids_to_provider(input_sample_path, output_path, output_sheet_name='Provider')
    # After creating Output.xlsx, copy NPI values
    from Name import copy_name_column
    copy_name_column(input_sample_path, output_path, output_sheet_name='Provider')
    from NPI import copy_npi_column
    copy_npi_column(input_sample_path, output_path, output_sheet_name='Provider')
    from Gender import copy_gender_column
    copy_gender_column(input_sample_path, output_path, output_sheet_name='Provider')
    from Professional_statement import copy_professional_statement
    copy_professional_statement(input_sample_path, output_path, output_sheet_name='Provider')
    from Langauges import copy_languages
    copy_languages(input_sample_path, output_path, output_sheet_name='Provider')
    from Headshot import copy_headshot_column
    copy_headshot_column(input_sample_path, output_path, output_sheet_name='Provider')
    from PatientsAccepted import copy_patients_accepted
    copy_patients_accepted(input_sample_path, output_path, output_sheet_name='Provider')
    from Professionalsuffix import copy_professional_suffix
    copy_professional_suffix(input_sample_path, output_path, output_sheet_name='Provider')
    # Add hospital affiliation dropdowns
    from Hospitalaff import main as add_hospital_affiliation_dropdowns
    add_hospital_affiliation_dropdowns()
    # Add provider type dropdowns
    from Providertype import add_provider_type_dropdown
    add_provider_type_dropdown()
    # Write formulas to Provider Type (Substatus) ID
    write_provider_type_substatus_id_formula(output_path, sheet_name='Provider')
    # Add Enterprise Scheduling Flag column with dropdown
    from EnterpriseSchedulingFlag import add_enterprise_scheduling_flag_column
    add_enterprise_scheduling_flag_column(output_path, sheet_name='Provider')
    # Add Board Certification 1-5 dropdowns
    add_board_certification_dropdowns(output_path, sheet_name='Provider')
    # Add Sub Board Certification 1-5 dropdowns
    add_sub_board_certification_dropdowns(output_path, sheet_name='Provider')
    # Write Board Cert ID 1 formula
    write_board_cert_id_1_formula(output_path, sheet_name='Provider')
    # Write Professional Suffix ID 1 formula
    write_professional_suffix_id_1_formula(output_path, sheet_name='Provider')
    # Add Location sheet processing
    from Locationsheet import process_location_sheet
    process_location_sheet(input_sample_path, output_path)
    # Open the output file automatically after all processing is complete
    # os.startfile(output_path)
    print("The data transposition is complete and is saved as 'Output.xlsx'")

# After Output.xlsx is created, run Specialtyapi.py
import subprocess
subprocess.run(["python", "Specialtyapi.py"], check=True)
