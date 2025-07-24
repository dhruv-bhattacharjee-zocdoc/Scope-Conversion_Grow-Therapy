import shutil
import openpyxl
from openpyxl.styles import PatternFill

# Define file paths
source_file = 'Excel Files/Output.xlsx'
destination_file = 'Excel Files/Mergedoutput.xlsx'
input_file = 'Excel Files/Grow Therapy - Sample Data.xlsx'
reference_file = 'Excel Files/json_to_excel.xlsx'

# Step 1: Copy the output file to create the merged output
shutil.copyfile(source_file, destination_file)

# Step 2: Open the input file and get the 'CORRECT Location' column
wb_input = openpyxl.load_workbook(input_file)
ws_input = wb_input.active
input_headers = [cell.value for cell in next(ws_input.iter_rows(min_row=1, max_row=1))]
try:
    correct_location_col_idx = input_headers.index('CORRECT Location') + 1  # openpyxl is 1-indexed
except ValueError:
    raise Exception("'CORRECT Location' column not found in input file.")

# Get all values from the 'CORRECT Location' column (excluding header)
correct_location_values = [row[correct_location_col_idx - 1].value for row in ws_input.iter_rows(min_row=2)]

# Step 3: Open the merged output file and insert the column as the first column in the 'Location' sheet
wb_out = openpyxl.load_workbook(destination_file)
ws_location = wb_out['Location']

# Insert a new column at position 1
ws_location.insert_cols(1)
ws_location.cell(row=1, column=1, value='CORRECT Location')

# Write the values
for idx, value in enumerate(correct_location_values, start=2):
    ws_location.cell(row=idx, column=1, value=value)

# Step 4: Open the reference file and get all relevant columns
wb_ref = openpyxl.load_workbook(reference_file)
ws_ref = wb_ref.active
ref_headers = [cell.value for cell in next(ws_ref.iter_rows(min_row=1, max_row=1))]

# Map reference columns to output columns
column_mappings = {
    'address_1': 'Address line 1',
    'address_2': 'Address line 2 (Office/Suite #)',
    'city': 'City',
    'location_id': 'Location Cloud ID',
    'name': 'Location Name',
    'practice_id': 'Practice Cloud ID',
    'state': 'State',
    'zip': 'ZIP Code',
    'virtual_visit_type': 'Virtual Visit Type',
}

# Get indices for reference columns
ref_col_indices = {}
for ref_col in ['monolith_location_id'] + list(column_mappings.keys()):
    try:
        ref_col_indices[ref_col] = ref_headers.index(ref_col) + 1
    except ValueError:
        raise Exception(f"'{ref_col}' column not found in reference file.")

# Build a mapping from monolith_location_id to all needed fields
monolith_to_data = {}
for row in ws_ref.iter_rows(min_row=2):
    monolith_id = row[ref_col_indices['monolith_location_id'] - 1].value
    if monolith_id is not None:
        monolith_to_data[monolith_id] = {k: row[ref_col_indices[k] - 1].value for k in column_mappings.keys()}

# Step 5: Highlight cells in the 'CORRECT Location' column if they match any monolith_location_id
# and copy mapped values to the Location sheet
highlight_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')  # Yellow
# Find the column indices for output columns
location_headers = [cell.value for cell in next(ws_location.iter_rows(min_row=1, max_row=1))]
output_col_indices = {}
for out_col in list(column_mappings.values()):
    try:
        output_col_indices[out_col] = location_headers.index(out_col) + 1
    except ValueError:
        raise Exception(f"'{out_col}' column not found in Location sheet of destination file.")

for idx, value in enumerate(correct_location_values, start=2):
    cell = ws_location.cell(row=idx, column=1)
    if value in monolith_to_data:
        cell.fill = highlight_fill
        for ref_col, out_col in column_mappings.items():
            ws_location.cell(row=idx, column=output_col_indices[out_col], value=monolith_to_data[value][ref_col])

# Step 6: Move 'CORRECT Location' column to the last position
max_col = ws_location.max_column
max_row = ws_location.max_row
# Read the column values
correct_location_col = [ws_location.cell(row=r, column=1).value for r in range(1, max_row + 1)]
# Delete the first column
ws_location.delete_cols(1)
# Insert at the end
ws_location.insert_cols(max_col)
for r, val in enumerate(correct_location_col, start=1):
    ws_location.cell(row=r, column=max_col, value=val)
# Update header
ws_location.cell(row=1, column=max_col, value='CORRECT Location')

# Step 7: Delete the 'CORRECT Location' column before saving
# Find the column index for 'CORRECT Location' (should be last column now)
location_headers = [cell.value for cell in next(ws_location.iter_rows(min_row=1, max_row=1))]
try:
    correct_location_col_idx = location_headers.index('CORRECT Location') + 1
    ws_location.delete_cols(correct_location_col_idx)
except ValueError:
    pass  # If not found, do nothing

# Step 7.1: Update 'Location ID 1' in the Provider sheet based on 'CORRECT Location' and Location sheet mapping
ws_provider = wb_out['Provider']

# Get headers for Provider and Location sheets
provider_headers = [cell.value for cell in next(ws_provider.iter_rows(min_row=1, max_row=1))]
location_headers = [cell.value for cell in next(ws_location.iter_rows(min_row=1, max_row=1))]

try:
    provider_correct_location_idx = provider_headers.index('CORRECT Location') + 1
except ValueError:
    provider_correct_location_idx = None
try:
    provider_location_id_1_idx = provider_headers.index('Location ID 1') + 1
except ValueError:
    provider_location_id_1_idx = None
try:
    location_correct_location_idx = location_headers.index('CORRECT Location') + 1
except ValueError:
    location_correct_location_idx = None
try:
    location_cloud_id_idx = location_headers.index('Location Cloud ID') + 1
except ValueError:
    location_cloud_id_idx = None

if all(idx is not None for idx in [provider_correct_location_idx, provider_location_id_1_idx, location_correct_location_idx, location_cloud_id_idx]):
    # Build a mapping from CORRECT Location to Location Cloud ID from the Location sheet
    correct_loc_to_cloud_id = {}
    for row in ws_location.iter_rows(min_row=2, max_row=ws_location.max_row):
        correct_loc = row[location_correct_location_idx - 1].value
        cloud_id = row[location_cloud_id_idx - 1].value
        if correct_loc is not None:
            correct_loc_to_cloud_id[correct_loc] = cloud_id
    # Fill Location ID 1 in Provider sheet
    for row in ws_provider.iter_rows(min_row=2, max_row=ws_provider.max_row):
        correct_loc = row[provider_correct_location_idx - 1].value
        if correct_loc is not None and correct_loc in correct_loc_to_cloud_id:
            ws_provider.cell(row=row[0].row, column=provider_location_id_1_idx, value=correct_loc_to_cloud_id[correct_loc])

# Step 7.2: Update 'Location ID 1' through 'Location ID 5' in Provider sheet using reference file mapping
# Build a mapping from monolith_location_id to location_id from the reference file
monolith_to_location_id = {}
for row in ws_ref.iter_rows(min_row=2):
    monolith_id = row[ref_col_indices['monolith_location_id'] - 1].value
    location_id = row[ref_col_indices['location_id'] - 1].value
    if monolith_id is not None:
        monolith_to_location_id[monolith_id] = location_id

# Find indices for 'Location ID 1' through 'Location ID 5' in Provider sheet
location_id_cols = []
for i in range(1, 6):
    col_name = f'Location ID {i}'
    try:
        idx = provider_headers.index(col_name) + 1
        location_id_cols.append(idx)
    except ValueError:
        pass  # If the column doesn't exist, skip

# For each row in Provider, update the values in 'Location ID 1' to 'Location ID 5' if a mapping exists
for row in ws_provider.iter_rows(min_row=2, max_row=ws_provider.max_row):
    for col_idx in location_id_cols:
        monolith_id = row[col_idx - 1].value
        if monolith_id in monolith_to_location_id:
            ws_provider.cell(row=row[0].row, column=col_idx, value=monolith_to_location_id[monolith_id])

# Copy 'Location ID 1-5' to 'Location 1-5' in Provider sheet
provider_headers = [cell.value for cell in next(ws_provider.iter_rows(min_row=1, max_row=1))]
location_id_indices = []
location_indices = []
for i in range(1, 6):
    try:
        location_id_indices.append(provider_headers.index(f'Location ID {i}') + 1)
        location_indices.append(provider_headers.index(f'Location {i}') + 1)
    except ValueError:
        pass  # If the column doesn't exist, skip
for row in ws_provider.iter_rows(min_row=2, max_row=ws_provider.max_row):
    for loc_id_idx, loc_idx in zip(location_id_indices, location_indices):
        value = row[loc_id_idx - 1].value
        ws_provider.cell(row=row[0].row, column=loc_idx, value=value)

# Step 7.3: Set 'Location Type' to 'Virtual' if 'Virtual Visit Type' is 'ThirdPartyVideoVisit', else leave empty
# Ensure 'Location Type' column exists, if not, add it at the end
location_headers = [cell.value for cell in next(ws_location.iter_rows(min_row=1, max_row=1))]
try:
    location_type_col_idx = location_headers.index('Location Type') + 1
except ValueError:
    # Add 'Location Type' as the last column
    location_type_col_idx = ws_location.max_column + 1
    ws_location.cell(row=1, column=location_type_col_idx, value='Location Type')

# Find the column index for 'Virtual Visit Type'
try:
    virtual_visit_type_col_idx = location_headers.index('Virtual Visit Type') + 1
except ValueError:
    virtual_visit_type_col_idx = None

if virtual_visit_type_col_idx is not None:
    for row in ws_location.iter_rows(min_row=2, max_row=ws_location.max_row):
        vvt = row[virtual_visit_type_col_idx - 1].value
        if vvt == 'ThirdPartyVideoVisit':
            ws_location.cell(row=row[0].row, column=location_type_col_idx, value='Virtual')
        else:
            ws_location.cell(row=row[0].row, column=location_type_col_idx, value=None)

# Step 7.4: Fill 'Practice Name' in Location sheet from 'software' column in reference file
# Ensure 'Practice Name' column exists, if not, add it at the end
location_headers = [cell.value for cell in next(ws_location.iter_rows(min_row=1, max_row=1))]
try:
    practice_name_col_idx = location_headers.index('Practice Name') + 1
except ValueError:
    practice_name_col_idx = ws_location.max_column + 1
    ws_location.cell(row=1, column=practice_name_col_idx, value='Practice Name')

# Find the column index for 'software' in reference file
try:
    software_col_idx = ref_headers.index('software') + 1
except ValueError:
    software_col_idx = None

if software_col_idx is not None:
    # Build a mapping from monolith_location_id to software
    monolith_to_software = {}
    for row in ws_ref.iter_rows(min_row=2):
        monolith_id = row[ref_col_indices['monolith_location_id'] - 1].value
        software = row[software_col_idx - 1].value
        if monolith_id is not None:
            monolith_to_software[monolith_id] = software
    # Fill Practice Name in Location sheet
    for idx, value in enumerate(correct_location_values, start=2):
        if value in monolith_to_software:
            ws_location.cell(row=idx, column=practice_name_col_idx, value=monolith_to_software[value])
        else:
            ws_location.cell(row=idx, column=practice_name_col_idx, value=None)

# Step 7.5: Set the same value for 'Practice Name' and 'Practice Cloud ID' in the Provider sheet as in the Location sheet, matching by 'Location ID 1'
# Get column indices in Provider sheet
try:
    provider_practice_name_idx = provider_headers.index('Practice Name') + 1
except ValueError:
    provider_practice_name_idx = ws_provider.max_column + 1
    ws_provider.cell(row=1, column=provider_practice_name_idx, value='Practice Name')
try:
    provider_practice_cloud_id_idx = provider_headers.index('Practice Cloud ID') + 1
except ValueError:
    provider_practice_cloud_id_idx = ws_provider.max_column + 1
    ws_provider.cell(row=1, column=provider_practice_cloud_id_idx, value='Practice Cloud ID')

# Get column indices in Location sheet
location_headers = [cell.value for cell in next(ws_location.iter_rows(min_row=1, max_row=1))]
try:
    location_location_id_idx = location_headers.index('Location Cloud ID') + 1
except ValueError:
    location_location_id_idx = None
try:
    location_practice_name_idx = location_headers.index('Practice Name') + 1
except ValueError:
    location_practice_name_idx = None
try:
    location_practice_cloud_id_idx = location_headers.index('Practice Cloud ID') + 1
except ValueError:
    location_practice_cloud_id_idx = None

if None not in (location_location_id_idx, location_practice_name_idx, location_practice_cloud_id_idx):
    # Build a mapping from Location Cloud ID to Practice Name and Practice Cloud ID
    locid_to_practice = {}
    for row in ws_location.iter_rows(min_row=2, max_row=ws_location.max_row):
        loc_id = row[location_location_id_idx - 1].value
        practice_name = row[location_practice_name_idx - 1].value
        practice_cloud_id = row[location_practice_cloud_id_idx - 1].value
        if loc_id is not None:
            locid_to_practice[loc_id] = (practice_name, practice_cloud_id)
    # Set values in Provider sheet
    try:
        provider_location_id_1_idx = provider_headers.index('Location ID 1') + 1
    except ValueError:
        provider_location_id_1_idx = None
    if provider_location_id_1_idx is not None:
        for row in ws_provider.iter_rows(min_row=2, max_row=ws_provider.max_row):
            loc_id = row[provider_location_id_1_idx - 1].value
            if loc_id in locid_to_practice:
                practice_name, practice_cloud_id = locid_to_practice[loc_id]
                ws_provider.cell(row=row[0].row, column=provider_practice_name_idx, value=practice_name)
                ws_provider.cell(row=row[0].row, column=provider_practice_cloud_id_idx, value=practice_cloud_id)
            else:
                ws_provider.cell(row=row[0].row, column=provider_practice_name_idx, value=None)
                ws_provider.cell(row=row[0].row, column=provider_practice_cloud_id_idx, value=None)

# Step 7.6: Map 'SPECIALTIES' from snowflake.xlsx to 'Specialty ID 1' in Provider sheet, matching by NPI
snowflake_file = 'Excel Files/snowflake.xlsx'
wb_snow = openpyxl.load_workbook(snowflake_file)
ws_snow = wb_snow.active
snow_headers = [cell.value for cell in next(ws_snow.iter_rows(min_row=1, max_row=1))]
try:
    snow_npi_idx = snow_headers.index('NPI') + 1
    snow_specialties_idx = snow_headers.index('SPECIALTIES') + 1
except ValueError as e:
    raise Exception(f"Column not found in snowflake.xlsx: {e}")
# Build NPI to SPECIALTIES mapping from snowflake.xlsx
npi_to_specialty = {}
for row in ws_snow.iter_rows(min_row=2, max_row=ws_snow.max_row):
    npi = row[snow_npi_idx - 1].value
    specialty = row[snow_specialties_idx - 1].value
    if npi is not None:
        npi_to_specialty[str(npi).strip()] = specialty
# Get column indices in Provider sheet
provider_headers = [cell.value for cell in next(ws_provider.iter_rows(min_row=1, max_row=1))]
try:
    provider_npi_idx = provider_headers.index('NPI Number') + 1
except ValueError:
    provider_npi_idx = None
try:
    provider_specialty_id_1_idx = provider_headers.index('Specialty ID 1') + 1
except ValueError:
    provider_specialty_id_1_idx = None
if provider_npi_idx is not None and provider_specialty_id_1_idx is not None:
    for row in ws_provider.iter_rows(min_row=2, max_row=ws_provider.max_row):
        npi = row[provider_npi_idx - 1].value
        if npi is not None:
            specialty = npi_to_specialty.get(str(npi).strip())
            if specialty is not None:
                ws_provider.cell(row=row[0].row, column=provider_specialty_id_1_idx, value=specialty)

# Step 7.7: Map 'Specialty 1' in Provider sheet using 'Specialty ID 1' and ValidationAndReference sheet
if 'ValidationAndReference' in wb_out.sheetnames:
    ws_val = wb_out['ValidationAndReference']
    val_headers = [cell.value for cell in next(ws_val.iter_rows(min_row=1, max_row=1))]
    try:
        val_specialty_id_idx = val_headers.index('Specialty ID') + 1
        val_specialty_name_idx = val_headers.index('Specialty Name') + 1
    except ValueError as e:
        raise Exception(f"Column not found in ValidationAndReference sheet: {e}")
    # Build mapping from Specialty ID to Specialty Name
    specialty_id_to_name = {}
    for row in ws_val.iter_rows(min_row=2, max_row=ws_val.max_row):
        sid = row[val_specialty_id_idx - 1].value
        sname = row[val_specialty_name_idx - 1].value
        if sid is not None:
            specialty_id_to_name[str(sid).strip()] = sname
    # Get column indices in Provider sheet
    provider_headers = [cell.value for cell in next(ws_provider.iter_rows(min_row=1, max_row=1))]
    try:
        provider_specialty_id_1_idx = provider_headers.index('Specialty ID 1') + 1
        provider_specialty_1_idx = provider_headers.index('Specialty 1') + 1
    except ValueError:
        provider_specialty_id_1_idx = None
        provider_specialty_1_idx = None
    if provider_specialty_id_1_idx is not None and provider_specialty_1_idx is not None:
        for row in ws_provider.iter_rows(min_row=2, max_row=ws_provider.max_row):
            sid = row[provider_specialty_id_1_idx - 1].value
            if sid is not None:
                sname = specialty_id_to_name.get(str(sid).strip())
                if sname is not None:
                    ws_provider.cell(row=row[0].row, column=provider_specialty_1_idx, value=sname)
        # Add data validation dropdown to 'Specialty 1' column
        from openpyxl.worksheet.datavalidation import DataValidation
        from openpyxl.utils import get_column_letter
        specialty_1_col_letter = get_column_letter(provider_specialty_1_idx)
        dv = DataValidation(type="list", formula1='=ValidationAndReference!$K:$K', allow_blank=True)
        dv.error = 'Select a value from the dropdown.'
        dv.errorTitle = 'Invalid Entry'
        dv.prompt = 'Please select a specialty from the list.'
        dv.promptTitle = 'Specialty 1'
        ws_provider.add_data_validation(dv)
        dv.add(f'{specialty_1_col_letter}2:{specialty_1_col_letter}{ws_provider.max_row}')

# Replace 'Location 1' in Provider with formula referencing Location sheet
location_headers = [cell.value for cell in next(ws_location.iter_rows(min_row=1, max_row=1))]
try:
    loc_cloud_id_idx = location_headers.index('Location Cloud ID') + 1
    loc_name_idx = location_headers.index('Location Name') + 1
    loc_city_idx = location_headers.index('City') + 1
    loc_state_idx = location_headers.index('State') + 1
except ValueError as e:
    loc_cloud_id_idx = None
    loc_name_idx = None
    loc_city_idx = None
    loc_state_idx = None
if None not in (loc_cloud_id_idx, loc_name_idx, loc_city_idx, loc_state_idx):
    # Build mapping from Location Cloud ID to row number in Location sheet
    locid_to_row = {}
    for row in ws_location.iter_rows(min_row=2, max_row=ws_location.max_row):
        loc_id = row[loc_cloud_id_idx - 1].value
        if loc_id is not None:
            locid_to_row[str(loc_id).strip()] = row[0].row
    # Get column index for 'Location 1' in Provider
    provider_headers = [cell.value for cell in next(ws_provider.iter_rows(min_row=1, max_row=1))]
    try:
        provider_location_1_idx = provider_headers.index('Location 1') + 1
    except ValueError:
        provider_location_1_idx = None
    if provider_location_1_idx is not None:
        for row in ws_provider.iter_rows(min_row=2, max_row=ws_provider.max_row):
            loc_id = row[provider_location_1_idx - 1].value
            if loc_id is not None and str(loc_id).strip() in locid_to_row:
                loc_row = locid_to_row[str(loc_id).strip()]
                # Columns: A=1, D=4, F=6, C=3 in Location sheet
                formula = f'=IF(Location!A{loc_row}<>"",CONCATENATE(Location!A{loc_row}," ",Location!D{loc_row}," ",Location!F{loc_row}," ","(",Location!C{loc_row},")"),"")'
                ws_provider.cell(row=row[0].row, column=provider_location_1_idx, value=formula)

    # Repeat the formula mapping for 'Location 2' through 'Location 5'
    for n in range(2, 6):
        col_name = f'Location {n}'
        try:
            provider_location_n_idx = provider_headers.index(col_name) + 1
        except ValueError:
            provider_location_n_idx = None
        if provider_location_n_idx is not None:
            for row in ws_provider.iter_rows(min_row=2, max_row=ws_provider.max_row):
                loc_id = row[provider_location_n_idx - 1].value
                if loc_id is not None and str(loc_id).strip() in locid_to_row:
                    loc_row = locid_to_row[str(loc_id).strip()]
                    formula = f'=IF(Location!A{loc_row}<>"",CONCATENATE(Location!A{loc_row}," ",Location!D{loc_row}," ",Location!F{loc_row}," ","(",Location!C{loc_row},")"),"")'
                    ws_provider.cell(row=row[0].row, column=provider_location_n_idx, value=formula)

    # Add dropdown data validation to 'Location 1' through 'Location 5' columns in Provider sheet
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.utils import get_column_letter
    for n in range(1, 6):
        col_name = f'Location {n}'
        try:
            col_idx = provider_headers.index(col_name) + 1
        except ValueError:
            continue
        col_letter = get_column_letter(col_idx)
        dv = DataValidation(type="list", formula1='=Location!$B$2:$B$242', allow_blank=True)
        dv.error = 'Select a value from the dropdown.'
        dv.errorTitle = 'Invalid Entry'
        dv.prompt = f'Please select a Location for {col_name}.'
        dv.promptTitle = col_name
        ws_provider.add_data_validation(dv)
        dv.add(f'{col_letter}2:{col_letter}{ws_provider.max_row}')

# Save the updated workbook
wb_out.save(destination_file)

print(f"Copied '{source_file}' to '{destination_file}', mapped and highlighted columns, and deleted 'CORRECT Location' column before saving.") 