import openpyxl
import os
from openpyxl.styles import PatternFill

def copy_npi_column(input_file, output_file, input_sheet_name=None, output_sheet_name=None):
    """
    Copy values from the 'NPI' column in the input Excel file to the 'NPI Number' column in the output Excel file.
    If sheet names are not provided, use the first sheet in each workbook.
    For the output sheet, fill empty cells in the 'NPI Number' column with a light red fill, but only for rows corresponding to the raw table.
    """
    wb_in = openpyxl.load_workbook(input_file)
    wb_out = openpyxl.load_workbook(output_file)

    ws_in = wb_in[input_sheet_name] if input_sheet_name else wb_in.worksheets[0]
    ws_out = wb_out[output_sheet_name] if output_sheet_name else wb_out.worksheets[0]

    # Find the column index for 'NPI' in input and 'NPI Number' in output
    in_header = [cell.value for cell in next(ws_in.iter_rows(min_row=1, max_row=1))]
    out_header = [cell.value for cell in next(ws_out.iter_rows(min_row=1, max_row=1))]

    try:
        npi_col_in = in_header.index('NPI') + 1  # 1-based index
        npi_col_out = out_header.index('NPI Number') + 1
    except ValueError as e:
        raise Exception(f"Column not found: {e}")

    # Copy values from input to output (starting from row 2)
    input_data_rows = 0
    for row_idx, row in enumerate(ws_in.iter_rows(min_row=2), start=2):
        npi_value = row[npi_col_in - 1].value
        ws_out.cell(row=row_idx, column=npi_col_out, value=npi_value)
        input_data_rows += 1

    # Light red fill for empty cells in 'NPI Number' column, only for rows in the raw table
    light_red_fill = PatternFill(start_color='FFFFC7CE', end_color='FFFFC7CE', fill_type='solid')
    for row_idx in range(2, 2 + input_data_rows):
        cell = ws_out.cell(row=row_idx, column=npi_col_out)
        if cell.value is None or cell.value == "":
            cell.fill = light_red_fill

    wb_out.save(output_file) 