import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

def add_enterprise_scheduling_flag_column(output_path, sheet_name='Provider'):
    # Load the workbook and the Provider sheet
    wb = openpyxl.load_workbook(output_path)
    ws = wb[sheet_name]

    # Find the last column
    max_col = ws.max_column
    max_row = ws.max_row

    # Check if the column already exists
    header_row = 1
    col_names = [ws.cell(row=header_row, column=col).value for col in range(1, max_col + 1)]
    if 'Enterprise Scheduling Flag' in col_names:
        flag_col_idx = col_names.index('Enterprise Scheduling Flag') + 1
    else:
        flag_col_idx = max_col + 1
        ws.cell(row=header_row, column=flag_col_idx, value='Enterprise Scheduling Flag')

    # Set default value 'No' and apply dropdown to all rows (except header)
    for row in range(2, max_row + 1):
        ws.cell(row=row, column=flag_col_idx, value='No')

    # Create data validation dropdown
    dv = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
    dv.error = 'Select Yes or No'
    dv.errorTitle = 'Invalid Input'
    dv.prompt = 'Please select Yes or No'
    dv.promptTitle = 'Enterprise Scheduling Flag'

    # Add the data validation to the column (excluding header)
    cell_range = f'{get_column_letter(flag_col_idx)}2:{get_column_letter(flag_col_idx)}{max_row}'
    dv.add(cell_range)
    ws.add_data_validation(dv)

    # Save the workbook
    wb.save(output_path)
    print(f"Enterprise Scheduling Flag column added with dropdown in '{sheet_name}' sheet of {output_path}")

if __name__ == '__main__':
    add_enterprise_scheduling_flag_column('Excel Files/Output.xlsx', 'Provider')
