import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

output_file = 'Excel Files/Output.xlsx'
provider_sheet_name = 'Provider'
validation_sheet_name = 'ValidationAndReference'
validation_range = 'ValidationAndReference!$Q$2:$Q$9'

# The first value in the validation range (Q2) will be set as default
FIRST_OPTION_CELL = 'Q2'


def add_provider_type_dropdown():
    wb = openpyxl.load_workbook(output_file)
    if validation_sheet_name not in wb.sheetnames:
        raise Exception(f"'{validation_sheet_name}' sheet not found in output file.")
    ws = wb[provider_sheet_name]
    validation_ws = wb[validation_sheet_name]

    # Check if Q2:Q9 are filled
    provider_types = [validation_ws[f'Q{row}'].value for row in range(2, 10)]
    if not all(provider_types):
        raise Exception(f"Some values in {validation_sheet_name}!Q2:Q9 are missing. Please check the reference data.")

    # Get the first option from the validation range
    first_option = validation_ws[FIRST_OPTION_CELL].value

    # Find the column index for 'Provider Type'
    header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    try:
        provider_type_col = header.index('Provider Type') + 1  # openpyxl is 1-indexed
    except ValueError:
        raise Exception("'Provider Type' column not found in header.")

    max_row = ws.max_row

    # Remove any existing data validations for the Provider Type column
    to_remove = []
    for dv in ws.data_validations.dataValidation:
        for sqref in dv.sqref:
            if get_column_letter(provider_type_col) in str(sqref):
                to_remove.append(dv)
                break
    for dv in to_remove:
        ws.data_validations.dataValidation.remove(dv)

    # Create the data validation object
    dv = DataValidation(type="list", formula1=f'={validation_range}', allow_blank=True)
    dv.error = 'Select a value from the dropdown.'
    dv.errorTitle = 'Invalid Entry'
    dv.prompt = 'Please select a provider type from the list.'
    dv.promptTitle = 'Provider Type'
    ws.add_data_validation(dv)

    col_letter = get_column_letter(provider_type_col)
    dv_range = f'{col_letter}2:{col_letter}{max_row}'
    dv.add(dv_range)

    # Set the default value (first option) for all rows (if cell is empty)
    for row in range(2, max_row + 1):
        cell = ws.cell(row=row, column=provider_type_col)
        if cell.value is None or str(cell.value).strip() == '':
            cell.value = first_option

    wb.save(output_file)
    print('Provider Type dropdown and default values added successfully.')

if __name__ == '__main__':
    add_provider_type_dropdown()
