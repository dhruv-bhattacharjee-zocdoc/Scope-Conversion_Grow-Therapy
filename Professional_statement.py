import openpyxl
import os

def copy_professional_statement(input_file, output_file, input_sheet_name=None, output_sheet_name='Provider'):
    """
    Copy values from the 'Latest Description' column in the input Excel file to the 'Professional Statement' column in the 'Provider' sheet of the output Excel file.
    """
    wb_in = openpyxl.load_workbook(input_file)
    wb_out = openpyxl.load_workbook(output_file)

    ws_in = wb_in[input_sheet_name] if input_sheet_name else wb_in.worksheets[0]
    ws_out = wb_out[output_sheet_name]

    # Find the column index for 'Latest Description' in input and 'Professional Statement' in output
    in_header = [cell.value for cell in next(ws_in.iter_rows(min_row=1, max_row=1))]
    out_header = [cell.value for cell in next(ws_out.iter_rows(min_row=1, max_row=1))]

    try:
        desc_col_in = in_header.index('Latest Description') + 1  # 1-based index
        prof_col_out = out_header.index('Professional Statement') + 1
    except ValueError as e:
        raise Exception(f"Column not found: {e}")

    # Copy values from input to output (starting from row 2)
    for row_idx, row in enumerate(ws_in.iter_rows(min_row=2), start=2):
        desc_value = row[desc_col_in - 1].value
        if desc_value is not None and desc_value != "":
            ws_out.cell(row=row_idx, column=prof_col_out, value=desc_value)

    wb_out.save(output_file)
