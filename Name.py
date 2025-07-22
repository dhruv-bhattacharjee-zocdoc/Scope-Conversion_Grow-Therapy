import openpyxl
import os
from openpyxl.styles import PatternFill

def copy_name_column(input_file, output_file, input_sheet_name=None, output_sheet_name='Provider'):
    """
    Copy and split the 'Name' column from the input Excel file into 'First Name' and 'Last Name' columns in the Provider sheet of the output file.
    If there is a middle name, include it with the first name.
    E.g., 'John Mike Helington' -> First Name: 'John Mike', Last Name: 'Helington'
    """
    wb_in = openpyxl.load_workbook(input_file)
    wb_out = openpyxl.load_workbook(output_file)

    ws_in = wb_in[input_sheet_name] if input_sheet_name else wb_in.worksheets[0]
    ws_out = wb_out[output_sheet_name]

    # Find the column indexes
    in_header = [cell.value for cell in next(ws_in.iter_rows(min_row=1, max_row=1))]
    out_header = [cell.value for cell in next(ws_out.iter_rows(min_row=1, max_row=1))]

    try:
        name_col_in = in_header.index('Name') + 1  # 1-based index
        first_name_col_out = out_header.index('First Name') + 1
        last_name_col_out = out_header.index('Last Name') + 1
    except ValueError as e:
        raise Exception(f"Column not found: {e}")

    # Copy and split values from input to output (starting from row 2)
    red_fill = PatternFill(start_color='FFFFC7CE', end_color='FFFFC7CE', fill_type='solid')
    for row_idx, row in enumerate(ws_in.iter_rows(min_row=2), start=2):
        name_value = row[name_col_in - 1].value
        first_name = last_name = ''
        if name_value is not None and str(name_value).strip() != "":
            name_parts = str(name_value).strip().split()
            if len(name_parts) == 1:
                first_name = name_parts[0]
                last_name = ''
            elif len(name_parts) == 2:
                first_name, last_name = name_parts
            else:
                first_name = ' '.join(name_parts[:-1])
                last_name = name_parts[-1]
        # Write values
        ws_out.cell(row=row_idx, column=first_name_col_out, value=first_name)
        ws_out.cell(row=row_idx, column=last_name_col_out, value=last_name)
        # Fill red if either is empty
        if not first_name:
            ws_out.cell(row=row_idx, column=first_name_col_out).fill = red_fill
        if not last_name:
            ws_out.cell(row=row_idx, column=last_name_col_out).fill = red_fill

    wb_out.save(output_file)
